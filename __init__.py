from grammar_macro import grammar
import openpyxl

if __name__ == '__main__':
    wb = openpyxl.load_workbook('WeeklySurvey_NeedCleaningUntil2022July31.xlsx')
    
    ws = wb.active
    spell = grammar()
    spell.open_grammar()

    for i in range(2, ws.max_row+1):
        str_before = ws.cell(row=i, column=5).value

        if str_before == 'NULL' or str_before == '':
            continue
        else:
            ws.cell(row=i, column=8).value = spell.spell_check(str_before)
            
            while True:
                if ws.cell(row=i, column=8).value == str_before:
                    break
                else:
                    str_before = ws.cell(row=i, column=8).value
                    ws.cell(row=i, column=8).value = spell.spell_check(str_before)

    wb.save('WeeklySurvey_NeedCleaningUntil2022July31.xlsx')